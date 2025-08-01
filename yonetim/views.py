from datetime import timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db.models import Sum, Count, Q, Case, When, Value, F
from django.db.models.fields import DecimalField
from django.core.paginator import Paginator
from django.utils import timezone
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import login
from django.contrib.auth.forms import AuthenticationForm
from django.urls import reverse

from .models import Sube, Personel, GelirGider, Mesai
from .forms import SubeForm, PersonelForm, GelirGiderForm, MesaiForm

def _get_month_date_range(request):
    today = timezone.now().date()
    try:
        year = int(request.GET.get("year", today.year))
        month = int(request.GET.get("month", today.month))
    except (ValueError, TypeError):
        year = today.year
        month = today.month

    start_date = today.replace(year=year, month=month, day=1)
    if month == 12:
        end_date = start_date.replace(year=year + 1, month=1)
    else:
        end_date = start_date.replace(month=month + 1)
    return start_date, end_date, year, month


@login_required
def ana_sayfa(request):
    if request.user.groups.filter(name="Şube Müdürü").exists():
        return redirect("yonetim:mesai_listesi")

    today = timezone.now().date()
    try:
        year = int(request.GET.get("year", today.year))
        month = int(request.GET.get("month", today.month))
    except (ValueError, TypeError):
        year = today.year
        month = today.month

    start_date = today.replace(year=year, month=month, day=1)
    if month == 12:
        end_date = start_date.replace(year=year + 1, month=1) - timedelta(days=1)
    else:
        end_date = start_date.replace(month=month + 1) - timedelta(days=1)

    # Optimize edilmiş sorgu
    sube_ozetleri = Sube.objects.annotate(
        aylik_gelir=Sum(
            Case(
                When(
                    gelirgider__tip="gelir",
                    gelirgider__tarih__range=[start_date, end_date],
                    then="gelirgider__tutar",
                ),
                default=Value(0),
                output_field=DecimalField(),
            )
        ),
        aylik_gider=Sum(
            Case(
                When(
                    gelirgider__tip="gider",
                    gelirgider__tarih__range=[start_date, end_date],
                    then="gelirgider__tutar",
                ),
                default=Value(0),
                output_field=DecimalField(),
            )
        ),
    ).annotate(aylik_net_kar=F("aylik_gelir") - F("aylik_gider"))

    # Genel toplamlar için ayrı sorgular (bunlar zaten verimli)
    toplam_gelir = (
        GelirGider.objects.filter(tip="gelir").aggregate(toplam=Sum("tutar"))["toplam"]
        or 0
    )
    toplam_gider = (
        GelirGider.objects.filter(tip="gider").aggregate(toplam=Sum("tutar"))["toplam"]
        or 0
    )
    net_kar = toplam_gelir - toplam_gider
    toplam_personel = Personel.objects.count()
    toplam_sube = Sube.objects.count()
    cafe_sube_sayisi = Sube.objects.filter(tur="cafe").count()
    otel_sube_sayisi = Sube.objects.filter(tur="otel").count()
    cafe_gelir = (
        GelirGider.objects.filter(tip="gelir", sube__tur="cafe").aggregate(
            toplam=Sum("tutar")
        )["toplam"]
        or 0
    )
    cafe_gider = (
        GelirGider.objects.filter(tip="gider", sube__tur="cafe").aggregate(
            toplam=Sum("tutar")
        )["toplam"]
        or 0
    )
    cafe_net = cafe_gelir - cafe_gider
    cafe_personel = Personel.objects.filter(sube__tur="cafe").count()
    otel_gelir = (
        GelirGider.objects.filter(tip="gelir", sube__tur="otel").aggregate(
            toplam=Sum("tutar")
        )["toplam"]
        or 0
    )
    otel_gider = (
        GelirGider.objects.filter(tip="gider", sube__tur="otel").aggregate(
            toplam=Sum("tutar")
        )["toplam"]
        or 0
    )
    otel_net = otel_gelir - otel_gider
    otel_personel = Personel.objects.filter(sube__tur="otel").count()

    prev_month_date = start_date - timedelta(days=1)
    next_month_date = end_date + timedelta(days=1)

    context = {
        "toplam_gelir": toplam_gelir,
        "toplam_gider": toplam_gider,
        "net_kar": net_kar,
        "toplam_personel": toplam_personel,
        "toplam_sube": toplam_sube,
        "cafe_sube_sayisi": cafe_sube_sayisi,
        "otel_sube_sayisi": otel_sube_sayisi,
        "cafe_gelir": cafe_gelir,
        "cafe_gider": cafe_gider,
        "cafe_net": cafe_net,
        "cafe_personel": cafe_personel,
        "otel_gelir": otel_gelir,
        "otel_gider": otel_gider,
        "otel_net": otel_net,
        "otel_personel": otel_personel,
        "sube_ozetleri": sube_ozetleri,
        "current_date": start_date,
        "prev_month": {"year": prev_month_date.year, "month": prev_month_date.month},
        "next_month": {"year": next_month_date.year, "month": next_month_date.month},
        "today": today,
    }
    return render(request, "yonetim/ana_sayfa.html", context)


@login_required
@permission_required("yonetim.view_sube", raise_exception=True)
def subeler_listesi(request):
    subeler_query = (
        Sube.objects.annotate(
            personel_sayisi=Count("personel", distinct=True),
            toplam_gelir=Sum(
                Case(
                    When(gelirgider__tip="gelir", then="gelirgider__tutar"),
                    default=Value(0),
                    output_field=DecimalField(),
                )
            ),
            toplam_gider=Sum(
                Case(
                    When(gelirgider__tip="gider", then="gelirgider__tutar"),
                    default=Value(0),
                    output_field=DecimalField(),
                )
            ),
        )
        .annotate(net_kar=F("toplam_gelir") - F("toplam_gider"))
        .order_by("-ad")
    )

    q = request.GET.get("q")
    if q:
        subeler_query = subeler_query.filter(ad__icontains=q)

    paginator = Paginator(subeler_query, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    context = {
        "page_obj": page_obj,
        "subeler": page_obj,
        "q": q,
    }
    return render(request, "yonetim/subeler_listesi.html", context)


@login_required
@permission_required("yonetim.add_sube", raise_exception=True)
def sube_ekle(request):
    if request.method == "POST":
        form = SubeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Şube başarıyla eklendi.")
            return redirect("yonetim:subeler_listesi")
    else:
        form = SubeForm()
    context = {"form": form}
    return render(request, "yonetim/sube_form.html", context)


@login_required
@permission_required("yonetim.change_sube", raise_exception=True)
def sube_duzenle(request, pk):
    sube = get_object_or_404(Sube, pk=pk)
    if request.method == "POST":
        form = SubeForm(request.POST, instance=sube)
        if form.is_valid():
            form.save()
            messages.success(request, "Şube başarıyla güncellendi.")
            return redirect("yonetim:subeler_listesi")
    else:
        form = SubeForm(instance=sube)
    context = {"form": form, "sube": sube}
    return render(request, "yonetim/sube_form.html", context)


@login_required
@permission_required("yonetim.delete_sube", raise_exception=True)
def sube_sil(request, pk):
    sube = get_object_or_404(Sube, pk=pk)
    if request.method == "POST":
        sube.delete()
        messages.success(request, "Şube başarıyla silindi.")
        return redirect("yonetim:subeler_listesi")
    context = {
        "item_name": f"Şube: {sube.ad}",
        "cancel_url": reverse("yonetim:subeler_listesi"),
    }
    return render(request, "yonetim/sil_onay.html", context)


@login_required
@permission_required("yonetim.view_personel", raise_exception=True)
def personel_listesi(request):
    personeller_query = (
        Personel.objects.select_related("sube")
        .annotate(toplam_mesai_saati=Sum("mesai__saat"))
        .order_by("ad", "soyad")
    )

    sube_id = request.GET.get("sube")
    if sube_id:
        personeller_query = personeller_query.filter(sube_id=sube_id)

    q = request.GET.get("q")
    if q:
        personeller_query = personeller_query.filter(
            Q(ad__icontains=q) | Q(soyad__icontains=q) | Q(pozisyon__icontains=q)
        )

    paginator = Paginator(personeller_query, 15)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "personeller": page_obj,
        "subeler": Sube.objects.all(),
        "q": q,
        "secili_sube": sube_id,
    }
    return render(request, "yonetim/personel_listesi.html", context)


@login_required
@permission_required("yonetim.add_personel", raise_exception=True)
def personel_ekle(request):
    if request.method == "POST":
        form = PersonelForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Personel başarıyla eklendi.")
            return redirect("yonetim:personel_listesi")
    else:
        form = PersonelForm()
    context = {"form": form}
    return render(request, "yonetim/personel_form.html", context)


@login_required
@permission_required("yonetim.change_personel", raise_exception=True)
def personel_duzenle(request, pk):
    personel = get_object_or_404(Personel, pk=pk)
    if request.method == "POST":
        form = PersonelForm(request.POST, instance=personel)
        if form.is_valid():
            form.save()
            messages.success(request, "Personel başarıyla güncellendi.")
            return redirect("yonetim:personel_listesi")
    else:
        form = PersonelForm(instance=personel)
    context = {"form": form, "personel": personel}
    return render(request, "yonetim/personel_form.html", context)


@login_required
@permission_required("yonetim.delete_personel", raise_exception=True)
def personel_sil(request, pk):
    personel = get_object_or_404(Personel, pk=pk)
    if request.method == "POST":
        personel.delete()
        messages.success(request, "Personel başarıyla silindi.")
        return redirect("yonetim:personel_listesi")
    context = {
        "item_name": f"Personel: {personel.tam_ad}",
        "cancel_url": reverse("yonetim:personel_listesi"),
    }
    return render(request, "yonetim/sil_onay.html", context)


@login_required
@permission_required("yonetim.view_mesai", raise_exception=True)
def mesai_listesi(request):
    start_date, end_date, year, month = _get_month_date_range(request)

    mesailer_query = Mesai.objects.select_related("personel", "personel__sube").filter(
        tarih__gte=start_date, tarih__lt=end_date
    )

    personel_id = request.GET.get("personel")
    if personel_id:
        mesailer_query = mesailer_query.filter(personel_id=personel_id)
    sube_id = request.GET.get("sube")
    if sube_id:
        mesailer_query = mesailer_query.filter(personel__sube_id=sube_id)
    q = request.GET.get("q")
    if q:
        mesailer_query = mesailer_query.filter(
            Q(personel__ad__icontains=q)
            | Q(personel__soyad__icontains=q)
            | Q(aciklama__icontains=q)
        )

    mesailer_query = mesailer_query.order_by("-tarih", "-saat")

    toplam_mesai = mesailer_query.aggregate(Sum("saat"))["saat__sum"] or 0
    personel_sayisi = mesailer_query.values("personel").distinct().count()
    ortalama_mesai = (toplam_mesai / personel_sayisi) if personel_sayisi > 0 else 0

    paginator = Paginator(mesailer_query, 15)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    prev_month_date = start_date - timedelta(days=1)
    next_month_date = end_date

    context = {
        "page_obj": page_obj,
        "mesailer": page_obj,
        "personeller": Personel.objects.all(),
        "subeler": Sube.objects.all(),
        "q": q,
        "secili_personel": personel_id,
        "secili_sube": sube_id,
        "current_date": start_date,
        "prev_month": {"year": prev_month_date.year, "month": prev_month_date.month},
        "next_month": {"year": next_month_date.year, "month": next_month_date.month},
        "toplam_mesai": toplam_mesai,
        "personel_sayisi": personel_sayisi,
        "ortalama_mesai": ortalama_mesai,
        "filtreler": {
            "year": year,
            "month": month,
            "personel_id": personel_id,
            "sube_id": sube_id,
            "q": q,
        }
    }
    return render(request, "yonetim/mesai_listesi.html", context)


@login_required
@permission_required("yonetim.add_mesai", raise_exception=True)
def mesai_ekle(request):
    if request.method == "POST":
        form = MesaiForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Mesai kaydı başarıyla eklendi.")
            return redirect("yonetim:mesai_listesi")
    else:
        form = MesaiForm()
    context = {"form": form}
    return render(request, "yonetim/mesai_form.html", context)


@login_required
@permission_required("yonetim.change_mesai", raise_exception=True)
def mesai_duzenle(request, pk):
    mesai = get_object_or_404(Mesai, pk=pk)
    if request.method == "POST":
        form = MesaiForm(request.POST, instance=mesai)
        if form.is_valid():
            form.save()
            messages.success(request, "Mesai kaydı başarıyla güncellendi.")
            return redirect("yonetim:mesai_listesi")
    else:
        form = MesaiForm(instance=mesai)
    context = {"form": form, "mesai": mesai}
    return render(request, "yonetim/mesai_form.html", context)


@login_required
@permission_required("yonetim.delete_mesai", raise_exception=True)
def mesai_sil(request, pk):
    mesai = get_object_or_404(Mesai, pk=pk)
    if request.method == "POST":
        mesai.delete()
        messages.success(request, "Mesai kaydı başarıyla silindi.")
        return redirect("yonetim:mesai_listesi")
    context = {
        "item_name": f"Mesai: {mesai.personel.tam_ad} - {mesai.tarih}",
        "cancel_url": reverse("yonetim:mesai_listesi"),
    }
    return render(request, "yonetim/sil_onay.html", context)


@login_required
@permission_required("yonetim.view_mesai", raise_exception=True)
def print_mesai_listesi(request):
    """Mesai listesinin yazıcı dostu versiyonunu hazırlar."""
    start_date, end_date, year, month = _get_month_date_range(request)

    mesailer_query = Mesai.objects.select_related("personel", "personel__sube").filter(
        tarih__gte=start_date, tarih__lt=end_date
    )

    personel_id = request.GET.get("personel")
    if personel_id:
        mesailer_query = mesailer_query.filter(personel_id=personel_id)
    sube_id = request.GET.get("sube")
    if sube_id:
        mesailer_query = mesailer_query.filter(personel__sube_id=sube_id)
    q = request.GET.get("q")
    if q:
        mesailer_query = mesailer_query.filter(
            Q(personel__ad__icontains=q)
            | Q(personel__soyad__icontains=q)
            | Q(aciklama__icontains=q)
        )

    mesailer_query = mesailer_query.order_by("-tarih", "-saat")

    context = {
        "mesailer": mesailer_query,
        "current_date": start_date,
    }
    return render(request, "yonetim/mesai_listesi_print.html", context)


@login_required
@permission_required("yonetim.view_personel", raise_exception=True)
def print_personel_listesi(request):
    """Personel listesinin yazıcı dostu versiyonunu hazırlar."""
    personeller_query = (
        Personel.objects.select_related("sube")
        .annotate(toplam_mesai_saati=Sum("mesai__saat"))
        .order_by("ad", "soyad")
    )

    sube_id = request.GET.get("sube")
    if sube_id:
        personeller_query = personeller_query.filter(sube_id=sube_id)

    q = request.GET.get("q")
    if q:
        personeller_query = personeller_query.filter(
            Q(ad__icontains=q) | Q(soyad__icontains=q) | Q(pozisyon__icontains=q)
        )
    context = {"personeller": personeller_query}
    return render(request, "yonetim/personel_listesi_print.html", context)


@login_required
@permission_required("yonetim.view_personel", raise_exception=True)
def export_personel_excel(request):
    """Personel listesini .xlsx olarak dışa aktarır."""
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="personel_listesi.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Personel Listesi"

    # Başlıkları yaz
    headers = [
        "Ad",
        "Soyad",
        "Pozisyon",
        "Şube",
        "İşe Başlama Tarihi",
        "Telefon",
        "Email",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header

    # Verileri filtrele ve yaz
    personeller = Personel.objects.select_related("sube").all()
    sube_id = request.GET.get("sube")
    if sube_id:
        personeller = personeller.filter(sube_id=sube_id)
    q = request.GET.get("q")
    if q:
        personeller = personeller.filter(
            Q(ad__icontains=q) | Q(soyad__icontains=q) | Q(pozisyon__icontains=q)
        )

    for row_num, personel in enumerate(personeller, 2):
        worksheet.cell(row=row_num, column=1).value = personel.ad
        worksheet.cell(row=row_num, column=2).value = personel.soyad
        worksheet.cell(row=row_num, column=3).value = personel.pozisyon
        worksheet.cell(row=row_num, column=4).value = personel.sube.ad
        worksheet.cell(row=row_num, column=5).value = personel.ise_baslama_tarihi
        worksheet.cell(row=row_num, column=6).value = personel.telefon
        worksheet.cell(row=row_num, column=7).value = personel.email

    workbook.save(response)
    return response


@login_required
@permission_required("yonetim.view_gelirgider", raise_exception=True)
def export_gelir_gider_excel(request):
    """Gelir/Gider listesini tek sayfalı bir .xlsx olarak dışa aktarır."""
    gelir_giderler_qs = GelirGider.objects.select_related("sube").all()
    tip = request.GET.get("tip")
    if tip:
        gelir_giderler_qs = gelir_giderler_qs.filter(tip=tip)
    sube_id = request.GET.get("sube")
    if sube_id:
        gelir_giderler_qs = gelir_giderler_qs.filter(sube_id=sube_id)
    baslangic = request.GET.get("baslangic")
    if baslangic:
        gelir_giderler_qs = gelir_giderler_qs.filter(tarih__gte=baslangic)
    bitis = request.GET.get("bitis")
    if bitis:
        gelir_giderler_qs = gelir_giderler_qs.filter(tarih__lte=bitis)

    if not gelir_giderler_qs.exists():
        messages.warning(request, "Dışa aktarılacak veri bulunamadı.")
        return redirect("yonetim:gelir_gider_listesi")

    df = pd.DataFrame.from_records(
        gelir_giderler_qs.values_list(
            "sube__ad", "tip", "kategori", "tutar", "tarih", "aciklama"
        ),
        columns=["Şube", "Tip", "Kategori", "Tutar", "Tarih", "Açıklama"],
    )
    df["Kategori"] = df["Kategori"].apply(
        lambda x: dict(GelirGider.KATEGORI_SECENEKLERI).get(x, x)
    )
    df["Tip"] = df["Tip"].apply(lambda x: dict(GelirGider.TIP_SECENEKLERI).get(x, x))

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Gelir_Gider_Detay_{timezone.now().strftime("%Y-%m-%d")}.xlsx'
    )

    with pd.ExcelWriter(response, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Gelir Gider Listesi", index=False)
        worksheet = writer.sheets["Gelir Gider Listesi"]
        header_font = Font(bold=True)
        for cell in worksheet["1:1"]:
            cell.font = header_font
        for col_idx, col in enumerate(df.columns, 1):
            worksheet.column_dimensions[get_column_letter(col_idx)].width = 20

    return response


@login_required
@permission_required("yonetim.view_mesai", raise_exception=True)
def export_mesai_excel(request):
    """Mesai listesini ve personel bazında özetini .xlsx olarak dışa aktarır."""
    today = timezone.now().date()
    try:
        year = int(request.GET.get("year", today.year))
        month = int(request.GET.get("month", today.month))
    except (ValueError, TypeError):
        year = today.year
        month = today.month
    start_date = today.replace(year=year, month=month, day=1)
    if month == 12:
        end_date = start_date.replace(year=year + 1, month=1)
    else:
        end_date = start_date.replace(month=month + 1)

    mesailer_qs = Mesai.objects.select_related("personel", "personel__sube").filter(
        tarih__gte=start_date, tarih__lt=end_date
    )

    personel_id = request.GET.get("personel")
    if personel_id:
        mesailer_qs = mesailer_qs.filter(personel_id=personel_id)
    sube_id = request.GET.get("sube")
    if sube_id:
        mesailer_qs = mesailer_qs.filter(personel__sube_id=sube_id)
    q = request.GET.get("q")
    if q:
        mesailer_qs = mesailer_qs.filter(
            Q(personel__ad__icontains=q)
            | Q(personel__soyad__icontains=q)
            | Q(aciklama__icontains=q)
        )

    if not mesailer_qs.exists():
        messages.warning(request, "Dışa aktarılacak mesai verisi bulunamadı.")
        return redirect("yonetim:mesai_listesi")

    df = pd.DataFrame.from_records(
        mesailer_qs.values(
            "personel__ad",
            "personel__soyad",
            "personel__sube__ad",
            "tarih",
            "saat",
            "aciklama",
        )
    )
    df["Personel"] = df["personel__ad"] + " " + df["personel__soyad"]
    df = df.rename(
        columns={
            "personel__sube__ad": "Şube",
            "tarih": "Tarih",
            "saat": "Saat",
            "aciklama": "Açıklama",
        }
    )
    df = df[["Personel", "Şube", "Tarih", "Saat", "Açıklama"]]

    personel_ozet = df.groupby("Personel")["Saat"].sum().reset_index()
    personel_ozet = personel_ozet.rename(columns={"Saat": "Toplam Mesai Saati"})
    personel_ozet = personel_ozet.sort_values(by="Toplam Mesai Saati", ascending=False)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Mesai_Raporu_{year}-{month}.xlsx"'
    )

    with pd.ExcelWriter(response, engine="openpyxl") as writer:
        personel_ozet.to_excel(writer, sheet_name="Personel Mesai Özeti", index=False)
        df.to_excel(writer, sheet_name="Tüm Mesai Kayıtları", index=False)

        ozet_sheet = writer.sheets["Personel Mesai Özeti"]
        detay_sheet = writer.sheets["Tüm Mesai Kayıtları"]

        header_font = Font(bold=True)
        for cell in ozet_sheet["1:1"]:
            cell.font = header_font
        for cell in detay_sheet["1:1"]:
            cell.font = header_font

        ozet_sheet.column_dimensions["A"].width = 30
        ozet_sheet.column_dimensions["B"].width = 25

        detay_sheet.column_dimensions["A"].width = 30
        detay_sheet.column_dimensions["B"].width = 20
        detay_sheet.column_dimensions["C"].width = 15
        detay_sheet.column_dimensions["D"].width = 10
        detay_sheet.column_dimensions["E"].width = 40

    return response


@login_required
def api_personel_by_sube(request):
    sube_id = request.GET.get("sube_id")
    personeller = Personel.objects.filter(sube_id=sube_id).order_by("ad")
    data = {"personeller": [{"id": p.id, "tam_ad": p.tam_ad} for p in personeller]}
    return JsonResponse(data)


def custom_login_view(request):
    if request.user.is_authenticated:
        if request.user.groups.filter(name="Şube Müdürü").exists():
            return redirect("yonetim:mesai_listesi")
        return redirect("yonetim:ana_sayfa")

    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            if user.groups.filter(name="Şube Müdürü").exists():
                return redirect("yonetim:mesai_listesi")
            if user.is_staff or user.is_superuser:
                return redirect("/admin/")
            return redirect("yonetim:ana_sayfa")
        else:
            messages.error(request, "Kullanıcı adı veya şifre hatalı.")
    else:
        form = AuthenticationForm()
    return render(request, "giris.html", {"form": form})


@login_required
@permission_required("yonetim.view_gelirgider", raise_exception=True)
def gelir_gider_listesi(request):
    gelir_giderler_query = GelirGider.objects.select_related("sube").all()

    tip = request.GET.get("tip")
    if tip:
        gelir_giderler_query = gelir_giderler_query.filter(tip=tip)
    sube_id = request.GET.get("sube")
    if sube_id:
        gelir_giderler_query = gelir_giderler_query.filter(sube_id=sube_id)
    baslangic = request.GET.get("baslangic")
    if baslangic:
        gelir_giderler_query = gelir_giderler_query.filter(tarih__gte=baslangic)
    bitis = request.GET.get("bitis")
    if bitis:
        gelir_giderler_query = gelir_giderler_query.filter(tarih__lte=bitis)

    gelir_giderler_query = gelir_giderler_query.order_by("-tarih", "-id")

    per_page = request.GET.get("per_page", 15)  # Default to 15 items per page
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 15

    paginator = Paginator(gelir_giderler_query, per_page)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    toplam_gelir = gelir_giderler_query.filter(tip="gelir").aggregate(toplam=Sum("tutar"))["toplam"] or 0
    toplam_gider = gelir_giderler_query.filter(tip="gider").aggregate(toplam=Sum("tutar"))["toplam"] or 0
    net_kar = toplam_gelir - toplam_gider

    context = {
        "page_obj": page_obj,
        "gelir_giderler": page_obj,
        "subeler": Sube.objects.all(),
        "secili_tip": tip,
        "secili_sube": sube_id,
        "baslangic": baslangic,
        "bitis": bitis,
        "per_page": per_page,
        "toplam_gelir": toplam_gelir,
        "toplam_gider": toplam_gider,
        "net_kar": net_kar,
    }
    return render(request, "yonetim/gelir_gider_listesi.html", context)


@login_required
@permission_required("yonetim.add_gelirgider", raise_exception=True)
def gelir_gider_ekle(request):
    if request.method == "POST":
        form = GelirGiderForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Gelir/Gider kaydı başarıyla eklendi.")
            return redirect("yonetim:gelir_gider_listesi")
    else:
        form = GelirGiderForm()
    context = {"form": form}
    return render(request, "yonetim/gelir_gider_form.html", context)


@login_required
@permission_required("yonetim.change_gelirgider", raise_exception=True)
def gelir_gider_duzenle(request, pk):
    gelir_gider = get_object_or_404(GelirGider, pk=pk)
    if request.method == "POST":
        form = GelirGiderForm(request.POST, instance=gelir_gider)
        if form.is_valid():
            form.save()
            messages.success(request, "Gelir/Gider kaydı başarıyla güncellendi.")
            return redirect("yonetim:gelir_gider_listesi")
    else:
        form = GelirGiderForm(instance=gelir_gider)
    context = {"form": form, "gelir_gider": gelir_gider}
    return render(request, "yonetim/gelir_gider_form.html", context)


@login_required
@permission_required("yonetim.delete_gelirgider", raise_exception=True)
def gelir_gider_sil(request, pk):
    gelir_gider = get_object_or_404(GelirGider, pk=pk)
    if request.method == "POST":
        gelir_gider.delete()
        messages.success(request, "Gelir/Gider kaydı başarıyla silindi.")
        return redirect("yonetim:gelir_gider_listesi")
    context = {
        "item_name": f"Gelir/Gider: {gelir_gider.tutar} ({gelir_gider.get_tip_display()})",
        "cancel_url": reverse("yonetim:gelir_gider_listesi"),
    }
    return render(request, "yonetim/sil_onay.html", context)


@login_required
@permission_required("yonetim.view_gelirgider", raise_exception=True)
def print_gelir_gider_listesi(request):
    """Gelir/Gider listesinin yazıcı dostu versiyonunu hazırlar."""
    gelir_giderler_query = GelirGider.objects.select_related("sube").all()

    tip = request.GET.get("tip")
    if tip:
        gelir_giderler_query = gelir_giderler_query.filter(tip=tip)
    sube_id = request.GET.get("sube")
    if sube_id:
        gelir_giderler_query = gelir_giderler_query.filter(sube_id=sube_id)
    baslangic = request.GET.get("baslangic")
    if baslangic:
        gelir_giderler_query = gelir_giderler_query.filter(tarih__gte=baslangic)
    bitis = request.GET.get("bitis")
    if bitis:
        gelir_giderler_query = gelir_giderler_query.filter(tarih__lte=bitis)

    gelir_giderler_query = gelir_giderler_query.order_by("-tarih", "-id")

    context = {
        "gelir_giderler": gelir_giderler_query,
        "baslangic": baslangic,
        "bitis": bitis,
    }
    return render(request, "yonetim/gelir_gider_listesi_print.html", context)

# Kullanılmayan API endpointleri kaldırıldı veya yetkilendirildi.